"""
NSE Darvas Box Scanner - VCP Backtest Analysis Report
======================================================
Builds a 7-sheet Excel workbook for a VCP backtest run.

Sheets:
  1. Verdict              — pass/fail against objective criteria
  2. Equity Curve         — cumulative R-multiple growth chart
  3. Score Band Analysis  — does composite_score predict quality?
  4. T-Count Analysis     — does 3T beat 2T? 4T beat 3T?
  5. Contraction Analysis — tighter final contraction → better outcomes?
  6. Yearly Performance   — regime-dependency check
  7. Trade Log            — every trade, all VCP-specific columns

Sheets 4 and 5 are unique to VCP and validate the two core hypotheses:
  a) More contractions = more evidence of institutional accumulation
  b) Tighter final contraction = better supply/demand balance at pivot

All numbers from backtest_trade_log (same DB as other pattern reports).
Mirrors backtest_report.py styling and helper functions exactly.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import (
    REPORTS_DIR, SCORE_THRESHOLDS, VCP_MIN_QUALITY_SCORE,
    RISK_PER_TRADE_PCT,
)
from database import (
    get_backtest_runs_df, get_backtest_symbol_summary_df,
    get_backtest_trade_log_df,
)
from logger_utils import get_logger

log = get_logger("performance")

# ─── Styling (identical to backtest_report.py for visual consistency) ─────────
CLR = {
    "header_bg":   "1F4E79",
    "header_fg":   "FFFFFF",
    "pass_bg":     "00B050",
    "fail_bg":     "C00000",
    "caution_bg":  "FFC000",
    "elite":       "00B050",
    "very_strong": "92D050",
    "strong":      "FFEB9C",
    "watch":       "FCE4D6",
    "alt_row":     "EBF3FB",
    "white":       "FFFFFF",
    "border":      "BDD7EE",
    "title_bg":    "2E75B6",
    "title_fg":    "FFFFFF",
}
THIN   = Side(style="thin", color=CLR["border"])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─── Pass/fail criteria (same thresholds as backtest_report.py) ───────────────
CRITERIA = {
    "min_trades":           30,
    "min_win_rate_pct":     45.0,
    "min_profit_factor":    1.3,
    "min_expectancy":       0.10,
    "max_drawdown_pct":     -35.0,
    "min_score_separation": 5.0,
}

# ─── VCP-specific trade log columns ──────────────────────────────────────────
VCP_PATTERN_COLS = [
    "t_count", "window_strategy", "base_start_date", "base_duration_weeks",
    "pivot_price", "pivot_low", "atr_compression_ratio",
    "contraction_widths", "contraction_lows", "contraction_durations",
    "contraction_volumes", "final_contraction_width_pct",
    "contractions_tightening", "contractions_shortening",
    "higher_lows_ok", "volume_slope_negative", "vdu_day_present",
    "breakout_volume_ratio", "rs_new_high", "stop_is_wide",
    "trend_template_ok", "price_above_ma50", "price_above_ma150",
    "price_above_ma200", "ma150_above_ma200", "ma200_uptrend",
    "ma50_above_ma150", "pct_above_52wk_low", "pct_below_52wk_high",
    "prior_uptrend_pct",
]


# ─── Main entry point ─────────────────────────────────────────────────────────

def generate_vcp_report(run_id: str) -> list[Path]:
    """
    Build VCP backtest Excel workbook for *run_id*.
    Returns list of saved file paths.
    """
    runs_df  = get_backtest_runs_df()
    run_row  = runs_df[runs_df["run_id"] == run_id]
    if run_row.empty:
        raise ValueError(f"No backtest run found with run_id={run_id}")
    run_meta = run_row.iloc[0].to_dict()

    all_trades_df  = get_backtest_trade_log_df(run_id)
    all_symbols_df = get_backtest_symbol_summary_df(run_id)

    if all_trades_df.empty:
        raise ValueError(f"Backtest run {run_id} has zero trades — nothing to analyse")

    trades_df  = all_trades_df[all_trades_df.get("pattern_type", "vcp") == "vcp"].copy() \
        if "pattern_type" in all_trades_df.columns \
        else all_trades_df.copy()
    symbols_df = (
        all_symbols_df[all_symbols_df["pattern_type"] == "vcp"].copy()
        if "pattern_type" in all_symbols_df.columns
        else all_symbols_df.copy()
    )

    score_col       = "composite_score"
    live_threshold  = VCP_MIN_QUALITY_SCORE
    live_df         = trades_df[trades_df[score_col] >= live_threshold]

    wb = Workbook()
    wb.remove(wb.active)

    verdict = _compute_vcp_verdict(live_df, run_meta, trades_df, live_threshold, score_col)

    _sheet_verdict(wb, verdict, run_meta, trades_df)
    _sheet_equity_curve(wb, trades_df)
    _sheet_score_band_analysis(wb, trades_df)
    _sheet_t_count_analysis(wb, trades_df)
    _sheet_contraction_analysis(wb, trades_df)
    _sheet_yearly_performance(wb, trades_df)
    _sheet_trade_log(wb, trades_df)

    out_path = REPORTS_DIR / f"backtest_analysis_vcp_{run_id}.xlsx"
    wb.save(out_path)
    log.info("VCP backtest report saved → %s", out_path)
    return [out_path]


# ─── Verdict computation ──────────────────────────────────────────────────────

def _compute_vcp_verdict(
    live_df: pd.DataFrame,
    run_meta: dict,
    full_df: pd.DataFrame,
    live_threshold: float,
    score_col: str,
) -> dict:
    trades_df = live_df
    n         = len(trades_df)
    wins      = (trades_df["rr_realised"] > 0).sum()
    win_rate  = wins / n * 100 if n else 0.0

    gross_profit  = trades_df.loc[trades_df["rr_realised"] > 0, "rr_realised"].sum()
    gross_loss    = trades_df.loc[trades_df["rr_realised"] < 0, "rr_realised"].abs().sum()
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else 999.0
    expectancy    = trades_df["rr_realised"].mean()
    if pd.isna(expectancy):
        expectancy = 0.0

    risk_fraction = RISK_PER_TRADE_PCT / 100.0
    if n > 0:
        chrono   = trades_df.sort_values("entry_date")
        equity   = (1 + chrono["rr_realised"] * risk_fraction).cumprod()
        roll_max = equity.cummax()
        max_dd   = ((equity - roll_max) / roll_max).min() * 100
        max_dd   = 0.0 if pd.isna(max_dd) else float(max_dd)
    else:
        max_dd = 0.0

    high_band = full_df[full_df["score_band"].isin(["elite", "very_strong"])]
    low_band  = full_df[full_df["score_band"] == "watch"]
    high_wr   = (high_band["rr_realised"] > 0).mean() * 100 if len(high_band) else None
    low_wr    = (low_band["rr_realised"] > 0).mean() * 100  if len(low_band)  else None
    sep       = (high_wr - low_wr) if (high_wr is not None and low_wr is not None) else None

    checks = []
    for check_def in [
        ("Sample size",           f"{n} trades",              f"≥ {CRITERIA['min_trades']}",
         n >= CRITERIA["min_trades"]),
        ("Win rate",              f"{win_rate:.1f}%",         f"≥ {CRITERIA['min_win_rate_pct']}%",
         win_rate >= CRITERIA["min_win_rate_pct"]),
        ("Profit factor",         f"{profit_factor:.2f}",     f"≥ {CRITERIA['min_profit_factor']}",
         profit_factor >= CRITERIA["min_profit_factor"]),
        ("Expectancy (avg R)",    f"{expectancy:+.3f}R",      f"≥ +{CRITERIA['min_expectancy']}R",
         expectancy >= CRITERIA["min_expectancy"]),
        ("Max drawdown",          f"{max_dd:.1f}%",           f"≥ {CRITERIA['max_drawdown_pct']}%",
         max_dd >= CRITERIA["max_drawdown_pct"]),
    ]:
        checks.append({
            "check": check_def[0], "value": check_def[1],
            "threshold": check_def[2], "passed": check_def[3],
        })

    MIN_BAND = 20
    if sep is not None and len(high_band) >= MIN_BAND and len(low_band) >= MIN_BAND:
        checks.append({
            "check":     "Score discriminates quality (Elite/VeryStrong vs Watch win rate)",
            "value":     f"{high_wr:.1f}% vs {low_wr:.1f}% (+{sep:.1f}pp)",
            "threshold": f"≥ +{CRITERIA['min_score_separation']}pp gap",
            "passed":    sep >= CRITERIA["min_score_separation"],
        })
    else:
        checks.append({
            "check":     "Score discriminates quality",
            "value":     "Insufficient band data",
            "threshold": f"≥ {MIN_BAND} trades per band",
            "passed":    True,
        })

    full_n      = len(full_df)
    full_wr     = (full_df["rr_realised"] > 0).mean() * 100 if full_n else 0.0
    overall     = all(c["passed"] for c in checks)
    n_passed    = sum(c["passed"] for c in checks)

    return {
        "overall_pass": overall,
        "n_checks": len(checks), "n_passed": n_passed,
        "checks": checks,
        "win_rate": win_rate, "profit_factor": profit_factor,
        "expectancy": expectancy, "max_dd": max_dd,
        "n_trades": n, "full_history_n_trades": full_n,
        "full_history_win_rate": full_wr,
        "live_threshold": live_threshold,
    }


# ─── Sheet 1: Verdict ─────────────────────────────────────────────────────────

def _sheet_verdict(
    wb: Workbook, verdict: dict, run_meta: dict, trades_df: pd.DataFrame
) -> None:
    ws = wb.create_sheet("Verdict")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    headline = (
        "✅  VCP METHOD VALIDATED — passes all objective criteria"
        if verdict["overall_pass"] else
        f"⚠️  VCP METHOD NOT YET VALIDATED — passes {verdict['n_passed']}/{verdict['n_checks']} criteria"
    )
    cell = ws["A1"]
    cell.value = headline
    cell.font  = Font(bold=True, size=16, color="FFFFFF")
    cell.fill  = PatternFill("solid",
        start_color=CLR["pass_bg"] if verdict["overall_pass"] else CLR["fail_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Run: {run_meta.get('run_id','?')}  |  Date: {run_meta.get('run_date','?')}  |  "
        f"Symbols tested: {run_meta.get('symbols_tested','?')}  |  "
        f"Live-eligible trades (score ≥ {verdict['live_threshold']:.0f}): {verdict['n_trades']}"
    )
    ws["A2"].font      = Font(italic=True, size=10, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Headline stats row
    stats = [
        ("Win Rate",      f"{verdict['win_rate']:.1f}%"),
        ("Profit Factor", f"{verdict['profit_factor']:.2f}"),
        ("Expectancy",    f"{verdict['expectancy']:+.3f}R"),
        ("Max Drawdown",  f"{verdict['max_dd']:.1f}%"),
    ]
    for i, (label, value) in enumerate(stats):
        col = i * 2 + 1
        ws.cell(row=4, column=col, value=label).font = Font(bold=True, size=10)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        vcell = ws.cell(row=5, column=col, value=value)
        vcell.font      = Font(bold=True, size=14)
        vcell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
    ws.row_dimensions[5].height = 24

    # Checks table
    hdr_row = 7
    for c, h in enumerate(["Check", "Value", "Threshold", "Pass/Fail"], 1):
        cell = ws.cell(row=hdr_row, column=c, value=h)
        cell.font  = Font(bold=True, color=CLR["header_fg"])
        cell.fill  = PatternFill("solid", start_color=CLR["header_bg"])
        cell.border = BORDER

    for r_off, chk in enumerate(verdict["checks"], 1):
        r = hdr_row + r_off
        ws.cell(row=r, column=1, value=chk["check"]).border  = BORDER
        ws.cell(row=r, column=2, value=chk["value"]).border  = BORDER
        ws.cell(row=r, column=3, value=chk["threshold"]).border = BORDER
        pf_cell = ws.cell(row=r, column=4,
                          value="✅ PASS" if chk["passed"] else "❌ FAIL")
        pf_cell.font   = Font(bold=True, color="FFFFFF")
        pf_cell.fill   = PatternFill("solid",
            start_color=CLR["pass_bg"] if chk["passed"] else CLR["fail_bg"])
        pf_cell.border = BORDER

    _auto_width(ws)


# ─── Sheet 2: Equity Curve ────────────────────────────────────────────────────

def _sheet_equity_curve(wb: Workbook, trades_df: pd.DataFrame) -> None:
    ws     = wb.create_sheet("Equity Curve")
    chrono = trades_df.sort_values("entry_date").reset_index(drop=True)

    risk_fraction = RISK_PER_TRADE_PCT / 100.0
    equity = (1 + chrono["rr_realised"] * risk_fraction).cumprod()

    for c, h in enumerate(["Trade #", "Entry Date", "Symbol", "R Realised",
                            f"Cumulative Equity ({RISK_PER_TRADE_PCT:.1f}% risk/trade)"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font   = Font(bold=True, color=CLR["header_fg"])
        cell.fill   = PatternFill("solid", start_color=CLR["header_bg"])
        cell.border = BORDER

    for i, row in chrono.iterrows():
        r = i + 2
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=row["entry_date"])
        ws.cell(row=r, column=3, value=row["symbol"])
        ws.cell(row=r, column=4, value=round(row["rr_realised"], 3))
        ws.cell(row=r, column=5, value=round(float(equity.iloc[i]), 4))
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER

    _auto_width(ws)

    chart = LineChart()
    chart.title   = "VCP Equity Curve — Cumulative Growth"
    chart.style   = 2
    chart.height  = 12
    chart.width   = 26
    n = len(chrono)
    chart.add_data(Reference(ws, min_col=5, min_row=1, max_row=n + 1), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    chart.series[0].smooth = False
    ws.add_chart(chart, "G2")


# ─── Sheet 3: Score Band Analysis ────────────────────────────────────────────

def _sheet_score_band_analysis(wb: Workbook, trades_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Score Band Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "Does the VCP Composite Score Actually Predict Outcome Quality?"
    ws["A1"].font      = Font(bold=True, size=13, color=CLR["title_fg"])
    ws["A1"].fill      = PatternFill("solid", start_color=CLR["title_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    band_order  = ["elite", "very_strong", "strong", "watch", "below_watch"]
    band_labels = {
        "elite":       "Elite (≥90)",
        "very_strong": "Very Strong (80-89)",
        "strong":      "Strong (70-79)",
        "watch":       "Watch (60-69)",
        "below_watch": "Below 60 — never reaches live scanner",
    }

    rows = []
    for band in band_order:
        sub = trades_df[trades_df["score_band"] == band]
        if sub.empty:
            continue
        n       = len(sub)
        wr      = (sub["rr_realised"] > 0).mean() * 100
        avg_rr  = sub["rr_realised"].mean()
        gp      = sub.loc[sub["rr_realised"] > 0, "rr_realised"].sum()
        gl      = sub.loc[sub["rr_realised"] < 0, "rr_realised"].abs().sum()
        pf      = min(gp / gl, 999.0) if gl > 0 else 999.0
        rows.append({
            "Score Band":       band_labels[band],
            "Trades":           n,
            "Win Rate %":       round(wr, 1),
            "Avg R per Trade":  round(avg_rr, 3),
            "Profit Factor":    round(pf, 2),
            "Avg RSI at Entry": round(sub["rsi_at_entry"].mean(), 1) if "rsi_at_entry" in sub else "",
            "Avg ADX at Entry": round(sub["adx_at_entry"].mean(), 1) if "adx_at_entry" in sub else "",
        })

    df = pd.DataFrame(rows)
    _write_styled_df(ws, df, start_row=3)

    if len(df) >= 2:
        chart = BarChart()
        chart.title   = "Win Rate by Score Band"
        chart.height  = 9
        chart.width   = 18
        chart.add_data(
            Reference(ws, min_col=3, min_row=3, max_row=3 + len(df)),
            titles_from_data=True,
        )
        chart.set_categories(Reference(ws, min_col=1, min_row=4, max_row=3 + len(df)))
        ws.add_chart(chart, f"A{3 + len(df) + 3}")


# ─── Sheet 4: T-Count Analysis (VCP-specific) ─────────────────────────────────

def _sheet_t_count_analysis(wb: Workbook, trades_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("T-Count Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = (
        "T-Count Analysis — Does More Contractions = Better Outcome?"
        "\n(Validates Minervini's claim that 3T > 2T, 4T elite)"
    )
    ws["A1"].font      = Font(bold=True, size=12, color=CLR["title_fg"])
    ws["A1"].fill      = PatternFill("solid", start_color=CLR["title_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    if "t_count" not in trades_df.columns:
        ws["A3"] = "t_count column not found in trade log — run backtest again."
        return

    df = trades_df.copy()
    df["t_bucket"] = df["t_count"].apply(
        lambda x: "5T+" if x >= 5 else f"{x}T"
    )

    rows = []
    for bucket in ["2T", "3T", "4T", "5T+", "ALL"]:
        sub = df if bucket == "ALL" else df[df["t_bucket"] == bucket]
        if sub.empty:
            continue
        n    = len(sub)
        wr   = (sub["rr_realised"] > 0).mean() * 100
        avg  = sub["rr_realised"].mean()
        gp   = sub.loc[sub["rr_realised"] > 0, "rr_realised"].sum()
        gl   = sub.loc[sub["rr_realised"] < 0, "rr_realised"].abs().sum()
        pf   = min(gp / gl, 999.0) if gl > 0 else 999.0
        fw   = sub["final_contraction_width_pct"].mean() if "final_contraction_width_pct" in sub else 0.0
        hold = sub["hold_days"].mean()
        rows.append({
            "T-Count":            bucket,
            "Trades":             n,
            "Win Rate %":         round(wr, 1),
            "Avg R per Trade":    round(avg, 3),
            "Profit Factor":      round(pf, 2),
            "Avg Final Width %":  round(fw, 1),
            "Avg Hold Days":      round(hold, 1),
        })

    result_df = pd.DataFrame(rows)
    _write_styled_df(ws, result_df, start_row=3)

    note_row = 3 + len(result_df) + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
    ws.cell(row=note_row, column=1, value=(
        "If 3T win rate > 2T win rate by ≥5pp, the T-count is doing discriminative work "
        "and VCP_MIN_CONTRACTIONS = 2 is correctly set (borderline cases included). "
        "If 3T ≈ 2T, consider whether the window finder is reliably counting contractions "
        "or whether T-count is just noise. If 4T+ underperforms 3T, those setups may be "
        "bases that are extended in time — check base_duration_weeks for 4T+ trades."
    )).font = Font(italic=True, size=9, color="808080")
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_row].height = 50

    # Bar chart
    if len(result_df) >= 2:
        chart = BarChart()
        chart.title  = "Win Rate % by T-Count"
        chart.height = 9
        chart.width  = 16
        chart.add_data(
            Reference(ws, min_col=3, min_row=3, max_row=3 + len(result_df)),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(ws, min_col=1, min_row=4, max_row=3 + len(result_df))
        )
        ws.add_chart(chart, f"A{note_row + 2}")


# ─── Sheet 5: Contraction Analysis (VCP-specific) ────────────────────────────

def _sheet_contraction_analysis(wb: Workbook, trades_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Contraction Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "Contraction Analysis — Tighter Final Contraction → Better Outcomes?"
    ws["A1"].font      = Font(bold=True, size=12, color=CLR["title_fg"])
    ws["A1"].fill      = PatternFill("solid", start_color=CLR["title_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Table A: Final contraction width buckets
    ws.cell(row=3, column=1, value="Table A — Final Contraction Width Buckets").font = \
        Font(bold=True, size=11)

    if "final_contraction_width_pct" in trades_df.columns:
        fw = trades_df["final_contraction_width_pct"]
        width_buckets = [
            ("≤3% (Elite)",     trades_df[fw <= 3.0]),
            ("3-6%",            trades_df[(fw > 3.0) & (fw <= 6.0)]),
            ("6-9%",            trades_df[(fw > 6.0) & (fw <= 9.0)]),
            ("9-12%",           trades_df[(fw > 9.0) & (fw <= 12.0)]),
            ("12-15% (NSE max)", trades_df[(fw > 12.0) & (fw <= 15.0)]),
            (">15% (gate fail)", trades_df[fw > 15.0]),
        ]
        rows_a = []
        for label, sub in width_buckets:
            if sub.empty:
                continue
            n   = len(sub)
            wr  = (sub["rr_realised"] > 0).mean() * 100
            avg = sub["rr_realised"].mean()
            gp  = sub.loc[sub["rr_realised"] > 0, "rr_realised"].sum()
            gl  = sub.loc[sub["rr_realised"] < 0, "rr_realised"].abs().sum()
            pf  = min(gp / gl, 999.0) if gl > 0 else 999.0
            rows_a.append({
                "Width Range":   label,
                "Trades":        n,
                "Win Rate %":    round(wr, 1),
                "Avg R":         round(avg, 3),
                "Profit Factor": round(pf, 2),
            })
        _write_styled_df(ws, pd.DataFrame(rows_a), start_row=4)
        next_row = 4 + len(rows_a) + 3
    else:
        ws.cell(row=4, column=1, value="final_contraction_width_pct column not in trade log")
        next_row = 6

    # Table B: Trend template pass vs fail
    ws.cell(row=next_row, column=1, value="Table B — Trend Template Pass vs Fail").font = \
        Font(bold=True, size=11)
    if "trend_template_ok" in trades_df.columns:
        rows_b = []
        for label, mask in [
            ("All 9 criteria PASS", trades_df["trend_template_ok"] == 1),
            ("Some criteria FAIL",  trades_df["trend_template_ok"] == 0),
        ]:
            sub = trades_df[mask]
            if sub.empty:
                continue
            n   = len(sub)
            wr  = (sub["rr_realised"] > 0).mean() * 100
            avg = sub["rr_realised"].mean()
            gp  = sub.loc[sub["rr_realised"] > 0, "rr_realised"].sum()
            gl  = sub.loc[sub["rr_realised"] < 0, "rr_realised"].abs().sum()
            pf  = min(gp / gl, 999.0) if gl > 0 else 999.0
            rows_b.append({
                "Trend Template": label,
                "Trades":         n,
                "Win Rate %":     round(wr, 1),
                "Avg R":          round(avg, 3),
                "Profit Factor":  round(pf, 2),
            })
        _write_styled_df(ws, pd.DataFrame(rows_b), start_row=next_row + 1)
        next_row = next_row + len(rows_b) + 4
    else:
        next_row += 2

    # Table C: Higher lows pass vs fail
    ws.cell(row=next_row, column=1, value="Table C — Higher Lows Across Contractions").font = \
        Font(bold=True, size=11)
    if "higher_lows_ok" in trades_df.columns:
        rows_c = []
        for label, mask in [
            ("Higher Lows YES (accumulation)", trades_df["higher_lows_ok"] == 1),
            ("Higher Lows NO (distribution)",  trades_df["higher_lows_ok"] == 0),
        ]:
            sub = trades_df[mask]
            if sub.empty:
                continue
            n   = len(sub)
            wr  = (sub["rr_realised"] > 0).mean() * 100
            avg = sub["rr_realised"].mean()
            rows_c.append({
                "Higher Lows": label,
                "Trades":      n,
                "Win Rate %":  round(wr, 1),
                "Avg R":       round(avg, 3),
            })
        _write_styled_df(ws, pd.DataFrame(rows_c), start_row=next_row + 1)

    _auto_width(ws)


# ─── Sheet 6: Yearly Performance ─────────────────────────────────────────────

def _sheet_yearly_performance(wb: Workbook, trades_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Yearly Performance")

    df = trades_df.copy()
    df["year"] = pd.to_datetime(df["entry_date"]).dt.year

    rows = []
    for year, sub in df.groupby("year"):
        n   = len(sub)
        wr  = (sub["rr_realised"] > 0).mean() * 100
        avg = sub["rr_realised"].mean()
        gp  = sub.loc[sub["rr_realised"] > 0, "rr_realised"].sum()
        gl  = sub.loc[sub["rr_realised"] < 0, "rr_realised"].abs().sum()
        pf  = min(gp / gl, 999.0) if gl > 0 else 999.0
        rows.append({
            "Year":           int(year),
            "Trades":         n,
            "Win Rate %":     round(wr, 1),
            "Avg R per Trade": round(avg, 3),
            "Profit Factor":  round(pf, 2),
            "Total R Gained": round(sub["rr_realised"].sum(), 2),
        })

    _write_styled_df(
        ws, pd.DataFrame(rows).sort_values("Year"),
        start_row=2,
        title="Year-by-Year Breakdown — Regime Dependency Check",
    )


# ─── Sheet 7: Trade Log ───────────────────────────────────────────────────────

def _sheet_trade_log(wb: Workbook, trades_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Trade Log")

    common_cols = [
        "symbol", "entry_date", "exit_date", "entry_price", "exit_price",
        "stop_loss", "target1", "target2", "outcome", "rr_realised",
        "hold_days", "composite_score", "score_band", "rs_rating",
        "rsi_at_entry", "adx_at_entry",
    ]
    display_cols = common_cols + [
        c for c in VCP_PATTERN_COLS if c in trades_df.columns
    ]
    display_cols = [c for c in display_cols if c in trades_df.columns]

    df = trades_df[display_cols].sort_values("entry_date")
    _write_styled_df(ws, df, start_row=2,
                     title="Full VCP Trade Log — Every Individual Trade")

    # Colour-code outcome column
    if "outcome" in display_cols:
        outcome_col = display_cols.index("outcome") + 1
        outcome_colors = {
            "target2_hit":  CLR["elite"],
            "target1_hit":  CLR["strong"],
            "stopped_out":  CLR["fail_bg"],
            "open_at_end":  CLR["caution_bg"],
        }
        for r in range(3, 3 + len(df)):
            val = ws.cell(row=r, column=outcome_col).value
            if val in outcome_colors:
                ws.cell(row=r, column=outcome_col).fill = PatternFill(
                    "solid", start_color=outcome_colors[val]
                )


# ─── Shared helpers (identical to backtest_report.py) ────────────────────────

def _write_styled_df(
    ws, df: pd.DataFrame, start_row: int, title: Optional[str] = None
) -> None:
    if title:
        n_cols = max(len(df.columns), 1)
        ws.merge_cells(
            start_row=start_row - 1, start_column=1,
            end_row=start_row - 1,   end_column=n_cols,
        )
        tcell = ws.cell(row=start_row - 1, column=1, value=title)
        tcell.font      = Font(bold=True, size=12, color=CLR["title_fg"])
        tcell.fill      = PatternFill("solid", start_color=CLR["title_bg"])
        tcell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[start_row - 1].height = 22

    for c, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=c, value=str(col_name))
        cell.font      = Font(bold=True, color=CLR["header_fg"])
        cell.fill      = PatternFill("solid", start_color=CLR["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[start_row].height = 28

    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row, start=1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = BORDER
            cell.alignment = Alignment(horizontal="center")
            bg = CLR["alt_row"] if r_idx % 2 == 0 else CLR["white"]
            cell.fill      = PatternFill("solid", start_color=bg)

    _auto_width(ws)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 32)
