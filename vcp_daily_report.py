"""
VCP Scanner — Daily Live Scan Excel Report  (v2)
=================================================
Generates a 3-sheet Excel workbook after every live scan:

  Sheet 1 — Signals    : confirmed signals (score ≥ VCP_MIN_QUALITY_SCORE)
  Sheet 2 — Watchlist  : forming patterns  (score ≥ VCP_WATCHLIST_MIN_SCORE)
  Sheet 3 — Scan Stats : run metadata, score distribution, regime breakdown

The report is ALWAYS generated regardless of --save flag.
Zero-signal scans still produce the Stats sheet so you can see why nothing passed.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR, VCP_MIN_QUALITY_SCORE, VCP_WATCHLIST_MIN_SCORE
from logger_utils import get_logger

log = get_logger("scanner")

# ─── Colours ─────────────────────────────────────────────────────────────────
CLR = {
    "header":       "1F4E79",
    "header_fg":    "FFFFFF",
    "breaking":     "00B050",
    "near":         "FFEB9C",
    "watching":     "FCE4D6",
    "watchlist":    "E2EFDA",
    "alt":          "EBF3FB",
    "white":        "FFFFFF",
    "border":       "BDD7EE",
    "title":        "2E75B6",
    "warn":         "FFC000",
    "fail":         "C00000",
    "pass":         "00B050",
}
THIN   = Side(style="thin", color=CLR["border"])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_BG = {
    "Breaking Out": CLR["breaking"],
    "Near Pivot":   CLR["near"],
    "Watching":     CLR["watching"],
}

SIGNAL_HEADERS = [
    ("Status",             14),
    ("Symbol",             14),
    ("Sector",             16),
    ("T-Count",             8),
    ("Score",               8),
    ("Timeframes",         12),
    ("Pivot Price",        12),
    ("Current",            10),
    ("Buy Zone Hi",        12),
    ("Stop Loss",          10),
    ("Target 1\n(+20%)",   12),
    ("Target 2\n(+40%)",   12),
    ("R:R",                 7),
    ("RS Rating",          10),
    ("Final\nWidth %",     10),
    ("ATR\nRatio",         10),
    ("SEPA\nTemplate",     12),
    ("VDU Day",             9),
    ("Higher\nLows",        9),
    ("Market\nRegime",     12),
    ("Vol\nCompress",      10),
    ("Position\nSize",     10),
    ("Capital\n₹",         12),
]

WATCHLIST_HEADERS = [
    ("Symbol",             14),
    ("Sector",             16),
    ("T-Count",             8),
    ("Score",               8),
    ("Timeframes",         12),
    ("Pivot Price",        12),
    ("Current",            10),
    ("% to Pivot",         10),
    ("Stop Loss",          10),
    ("ATR Ratio",          10),
    ("Final Width%",       11),
    ("Market Regime",      12),
    ("Why Below Gate",     28),
]


def _hdr_cell(ws, row: int, col: int, text: str, width: int | None = None) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, color=CLR["header_fg"])
    c.fill      = PatternFill("solid", start_color=CLR["header"])
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _title_row(ws, text: str, ncols: int, row: int = 1, bg: str | None = None) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=bg or CLR["header"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26


def _data_row(ws, row: int, values: list, status: str = "") -> None:
    for c, val in enumerate(values, 1):
        cell           = ws.cell(row=row, column=c, value=val)
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        bg = CLR["alt"] if row % 2 == 0 else CLR["white"]
        cell.fill      = PatternFill("solid", start_color=bg)

    # Override first-column status colour
    if status and status in STATUS_BG:
        ws.cell(row=row, column=1).fill = PatternFill("solid", start_color=STATUS_BG[status])
        ws.cell(row=row, column=1).font = Font(bold=True)


# ─── Sheet 1: Signals ────────────────────────────────────────────────────────

def _sheet_signals(wb: Workbook, signals: list, scan_date: date) -> None:
    ws = wb.create_sheet("Signals")
    ws.sheet_view.showGridLines = False

    n_breaking = sum(1 for s in signals if s.status == "Breaking Out")
    n_near     = sum(1 for s in signals if s.status == "Near Pivot")
    n_watch    = sum(1 for s in signals if s.status == "Watching")

    title = (
        f"VCP Live Signals — {scan_date}   |   {len(signals)} signals   |   "
        f"🚀 {n_breaking} Breaking Out   👀 {n_near} Near Pivot   📊 {n_watch} Watching"
        + ("   |   ⚠️ Market is sideways — scores reduced by regime multiplier" if n_breaking == 0 else "")
    )
    _title_row(ws, title, len(SIGNAL_HEADERS), row=1,
               bg=CLR["breaking"] if n_breaking > 0 else CLR["warn"])
    ws.row_dimensions[1].height = 28

    for c, (hdr, w) in enumerate(SIGNAL_HEADERS, 1):
        _hdr_cell(ws, 2, c, hdr, width=w)
    ws.row_dimensions[2].height = 32

    order = {"Breaking Out": 0, "Near Pivot": 1, "Watching": 2}
    ordered = sorted(signals, key=lambda s: (order.get(s.status, 3), -s.quality_score))

    for i, s in enumerate(ordered):
        r = 3 + i
        tfs = "+".join(t[0].upper() for t in (s.active_timeframes or []))
        pct_to_pivot = round((s.pivot_price / s.current_price - 1) * 100, 1) if s.current_price > 0 else 0
        _data_row(ws, r, [
            s.status,
            s.symbol.replace(".NS", ""),
            s.sector,
            f"{s.t_count}T",
            round(s.quality_score, 1),
            tfs or "D",
            s.pivot_price,
            s.current_price,
            s.buy_zone_high,
            s.stop_loss,
            s.target1,
            s.target2,
            s.rr_ratio,
            round(s.rs_rating, 0),
            round(s.final_contraction_width, 1),
            round(s.atr_compression_ratio, 2),
            "✓" if s.trend_template_ok else "✗",
            "✓" if s.vdu_day_present  else "✗",
            "✓" if s.higher_lows_ok   else "✗",
            s.market_regime.title(),
            round(getattr(s, "vol_compression_score", 0), 2),
            s.position_size,
            f"₹{s.capital_required:,.0f}",
        ], status=s.status)

    ws.freeze_panes = "A3"

    if not signals:
        ws.merge_cells("A3:M3")
        c = ws.cell(row=3, column=1,
                    value="No signals above threshold today — check Watchlist sheet for forming patterns")
        c.font      = Font(italic=True, color="808080")
        c.alignment = Alignment(horizontal="center")


# ─── Sheet 2: Watchlist ───────────────────────────────────────────────────────

def _sheet_watchlist(wb: Workbook, watchlist: list, scan_date: date) -> None:
    ws = wb.create_sheet("Watchlist")
    ws.sheet_view.showGridLines = False

    _title_row(
        ws,
        f"VCP Watchlist — {scan_date}   |   {len(watchlist)} forming patterns "
        f"(score {VCP_WATCHLIST_MIN_SCORE:.0f}–{VCP_MIN_QUALITY_SCORE:.0f})   |   "
        "These are setting up but not yet ready to buy",
        len(WATCHLIST_HEADERS), row=1, bg=CLR["title"],
    )
    ws.row_dimensions[1].height = 28

    for c, (hdr, w) in enumerate(WATCHLIST_HEADERS, 1):
        _hdr_cell(ws, 2, c, hdr, width=w)
    ws.row_dimensions[2].height = 28

    ordered = sorted(watchlist, key=lambda s: -s.quality_score)

    for i, s in enumerate(ordered):
        r = 3 + i
        tfs = "+".join(t[0].upper() for t in (s.active_timeframes or []))
        pct = round((s.pivot_price / s.current_price - 1) * 100, 1) if s.current_price > 0 else 0

        # Why below gate — show the main missing factor
        fails = [k for k, v in s.trend_template_details.items() if not v] if s.trend_template_details else []
        why = ""
        if not s.trend_template_ok and fails:
            why = f"SEPA fails: {', '.join(fails[:3])}"
        elif s.quality_score < VCP_MIN_QUALITY_SCORE:
            gap = VCP_MIN_QUALITY_SCORE - s.quality_score
            why = f"Score {s.quality_score:.1f} — {gap:.1f}pts below gate"

        _data_row(ws, r, [
            s.symbol.replace(".NS", ""),
            s.sector,
            f"{s.t_count}T",
            round(s.quality_score, 1),
            tfs or "D",
            s.pivot_price,
            s.current_price,
            f"{pct:+.1f}%",
            s.stop_loss,
            round(s.atr_compression_ratio, 2),
            round(s.final_contraction_width, 1),
            s.market_regime.title(),
            why,
        ])
        # Watchlist row colour
        ws.cell(row=r, column=1).fill = PatternFill("solid", start_color=CLR["watchlist"])

    ws.freeze_panes = "A3"

    if not watchlist:
        ws.merge_cells("A3:M3")
        c = ws.cell(row=3, column=1,
                    value="No watchlist candidates today — market may be extended or all stocks lack compression")
        c.font      = Font(italic=True, color="808080")
        c.alignment = Alignment(horizontal="center")


# ─── Sheet 3: Scan Stats ─────────────────────────────────────────────────────

def _sheet_stats(wb: Workbook, signals: list, watchlist: list,
                 all_vcps: list, scan_meta: dict, scan_date: date) -> None:
    ws = wb.create_sheet("Scan Stats")
    ws.sheet_view.showGridLines = False

    _title_row(ws, f"VCP Scan Statistics — {scan_date}", 4, row=1)

    # Key metrics table
    rows = [
        ("Symbols scanned",       scan_meta.get("symbols_scanned", "?")),
        ("Symbols with data",     scan_meta.get("symbols_with_data", "?")),
        ("Patterns detected",     scan_meta.get("patterns_detected", 0)),
        ("Above gate (Signals)",  len(signals)),
        (f"Forming (Watchlist ≥{VCP_WATCHLIST_MIN_SCORE:.0f})", len(watchlist)),
        ("Scan duration (secs)",  scan_meta.get("elapsed_secs", "?")),
        ("Market regime",         scan_meta.get("market_regime", "unknown").title()),
        ("Benchmark",             scan_meta.get("benchmark", "^NSEI")),
    ]
    for i, (label, val) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)

    # Score distribution
    if all_vcps:
        ws.cell(row=12, column=1, value="Score Distribution").font = Font(bold=True, size=11)
        buckets = [
            ("≥ 90 (Elite)",       [s for s in all_vcps if s.quality_score >= 90]),
            ("80–89 (Excellent)",  [s for s in all_vcps if 80 <= s.quality_score < 90]),
            ("70–79 (Good)",       [s for s in all_vcps if 70 <= s.quality_score < 80]),
            ("60–69 (Signals)",    [s for s in all_vcps if 60 <= s.quality_score < 70]),
            ("55–59 (Gate)",       [s for s in all_vcps if 55 <= s.quality_score < 60]),
            ("40–54 (Watchlist)",  [s for s in all_vcps if 40 <= s.quality_score < 55]),
            ("< 40 (Rejected)",    [s for s in all_vcps if s.quality_score < 40]),
        ]
        for j, (label, grp) in enumerate(buckets):
            r = 13 + j
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=len(grp))

    # Timeframe breakdown
    ws.cell(row=21, column=1, value="Timeframe Combos").font = Font(bold=True, size=11)
    from collections import Counter
    tf_counts = Counter(
        "+".join(sorted(s.active_timeframes or ["daily"]))
        for s in all_vcps
    )
    for j, (tf, count) in enumerate(tf_counts.most_common()):
        ws.cell(row=22 + j, column=1, value=tf)
        ws.cell(row=22 + j, column=2, value=count)

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)


# ─── Main Entry ──────────────────────────────────────────────────────────────

def generate_daily_vcp_report(
    signals:     list,
    watchlist:   Optional[list]  = None,
    all_vcps:    Optional[list]  = None,
    scan_meta:   Optional[dict]  = None,
    scan_date:   Optional[date]  = None,
) -> Optional[Path]:
    """
    Build the daily Excel report.

    signals   — VCPSignal objects with score ≥ VCP_MIN_QUALITY_SCORE
    watchlist — VCPSignal objects with score ≥ VCP_WATCHLIST_MIN_SCORE and < MIN
    all_vcps  — all detected VCPs (for stats sheet)
    scan_meta — dict with elapsed_secs, symbols_scanned, market_regime etc.
    """
    scan_date  = scan_date  or date.today()
    watchlist  = watchlist  or []
    all_vcps   = all_vcps   or signals + watchlist
    scan_meta  = scan_meta  or {}

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    _sheet_signals (wb, signals,  scan_date)
    _sheet_watchlist(wb, watchlist, scan_date)
    _sheet_stats   (wb, signals, watchlist, all_vcps, scan_meta, scan_date)

    out = REPORTS_DIR / f"vcp_daily_scan_{scan_date.isoformat()}.xlsx"
    wb.save(out)
    log.info(
        "Daily VCP report saved → %s  (%d signals, %d watchlist)",
        out, len(signals), len(watchlist),
    )
    return out
